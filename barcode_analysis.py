import openpyxl
import re
from itertools import combinations

# Load workbook
wb = openpyxl.load_workbook(r'2026-03-13-Barcode-and-Primers.xlsx')
ws = wb['Barcode Oligos']

# Extract barcode sequences from oligos
# Pattern: ...CCGATCT[BARCODE]NNN...
barcodes = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    well = row[0].value
    bc_name = row[1].value
    oligo = row[2].value
    if oligo is None:
        continue
    # Remove modification markers (* characters)
    clean = oligo.replace('*', '')
    # Extract barcode between CCGATCT and NNNNNNNN
    match = re.search(r'CCGATCT([ACGT]+?)N{8}', clean)
    if match:
        barcode = match.group(1)[:8]  # Use only the first 8 nucleotides
        barcodes[bc_name] = {'well': well, 'barcode': barcode}

print(f"Extracted {len(barcodes)} barcodes\n")

# --- Show all barcodes ---
print(f"{'Name':<8} {'Well':<6} {'Barcode':<12} {'Len':>3}  {'GC%':>5}")
print("-" * 42)
for name, info in barcodes.items():
    bc = info['barcode']
    gc = (bc.count('G') + bc.count('C')) / len(bc) * 100
    print(f"{name:<8} {info['well']:<6} {bc:<12} {len(bc):>3}  {gc:>5.1f}")

# --- GC Content Analysis ---
print("\n" + "=" * 60)
print("GC CONTENT ANALYSIS")
print("=" * 60)
print("Recommended GC content for barcodes: 25-75% (ideal: 40-60%)")
print("(Thermo Fisher recommends 40-60% for oligos;")
print(" PLOS ONE barcode study penalizes GC outside 40-60%)\n")

flagged_gc = []
for name, info in barcodes.items():
    bc = info['barcode']
    gc = (bc.count('G') + bc.count('C')) / len(bc) * 100
    status = ""
    if gc < 25 or gc > 75:
        status = "** POOR **"
        flagged_gc.append((name, bc, gc, "POOR"))
    elif gc < 37.5 or gc > 62.5:
        status = "* MARGINAL *"
        flagged_gc.append((name, bc, gc, "MARGINAL"))

if flagged_gc:
    print(f"{'Name':<8} {'Barcode':<12} {'GC%':>5}  {'Status'}")
    print("-" * 40)
    for name, bc, gc, status in flagged_gc:
        print(f"{name:<8} {bc:<12} {gc:>5.1f}  {status}")
else:
    print("All barcodes have acceptable GC content (40-60%).")

gc_values = []
for info in barcodes.values():
    bc = info['barcode']
    gc_values.append((bc.count('G') + bc.count('C')) / len(bc) * 100)
print(f"\nGC content range: {min(gc_values):.1f}% - {max(gc_values):.1f}%")
print(f"Mean GC content: {sum(gc_values)/len(gc_values):.1f}%")

# --- Hamming Distance Analysis ---
print("\n" + "=" * 60)
print("HAMMING DISTANCE ANALYSIS")
print("=" * 60)

def hamming_distance(s1, s2):
    if len(s1) != len(s2):
        return None  # Can't compute for different lengths
    return sum(c1 != c2 for c1, c2 in zip(s1, s2))

bc_names = list(barcodes.keys())
bc_seqs = [barcodes[n]['barcode'] for n in bc_names]

# Check if all barcodes are the same length
lengths = set(len(s) for s in bc_seqs)
print(f"Barcode lengths present: {lengths}")

# Compute all pairwise hamming distances
all_distances = []
low_hd_pairs = []
min_hd = float('inf')
min_pair = None

for i, j in combinations(range(len(bc_names)), 2):
    s1, s2 = bc_seqs[i], bc_seqs[j]
    if len(s1) != len(s2):
        continue
    hd = hamming_distance(s1, s2)
    all_distances.append(hd)
    if hd < min_hd:
        min_hd = hd
        min_pair = (bc_names[i], bc_names[j], s1, s2)
    if hd < 3:  # Flag pairs with HD < 3 (risk of barcode misassignment)
        low_hd_pairs.append((bc_names[i], bc_names[j], s1, s2, hd))

print(f"\nTotal pairwise comparisons: {len(all_distances)}")
if all_distances:
    print(f"Minimum Hamming distance: {min_hd}")
    print(f"  Between: {min_pair[0]} ({min_pair[2]}) and {min_pair[1]} ({min_pair[3]})")
    print(f"Maximum Hamming distance: {max(all_distances)}")
    print(f"Mean Hamming distance: {sum(all_distances)/len(all_distances):.2f}")

    # Distribution
    from collections import Counter
    dist = Counter(all_distances)
    print(f"\nHamming distance distribution:")
    for hd_val in sorted(dist.keys()):
        bar = "#" * min(dist[hd_val], 80)
        print(f"  HD={hd_val}: {dist[hd_val]:>5} pairs  {bar}")

    # Flag problematic pairs
    print(f"\n--- Pairs with Hamming distance < 3 (HIGH RISK of misassignment) ---")
    if low_hd_pairs:
        print(f"  {'HD':<4} {'BC1':<8} {'Seq1':<10} {'Well1':<6} {'BC2':<8} {'Seq2':<10} {'Well2':<6}")
        print(f"  {'-'*58}")
        for n1, n2, s1, s2, hd in sorted(low_hd_pairs, key=lambda x: x[4]):
            w1 = barcodes[n1]['well']
            w2 = barcodes[n2]['well']
            diff = ''.join('^' if a != b else ' ' for a, b in zip(s1, s2))
            print(f"  {hd:<4} {n1:<8} {s1:<10} {w1:<6} {n2:<8} {s2:<10} {w2:<6}")
            print(f"       {'':8} {diff}")
    else:
        print("  None found - all pairs have HD >= 3.")

    # Recommended minimum
    print(f"\n--- Recommendation ---")
    print(f"For 8-mer barcodes, a minimum Hamming distance of >= 3 is recommended")
    print(f"to tolerate up to 1 sequencing error (HD >= 2d+1 where d=errors to correct).")
    if min_hd >= 3:
        print(f">> Your barcode set PASSES with minimum HD = {min_hd}.")
    else:
        print(f">> WARNING: Your barcode set has minimum HD = {min_hd}, which is below the recommended threshold.")
